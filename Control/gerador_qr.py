import qrcode
import time
import os
from datetime import datetime
from urllib.parse import urlparse
#biblioteca feita para separar em partes o link onde o computador não conseguiria ler,separando em:
#scheme:https
#netloc:www.youtube.com
#path:watch
#query:v=abc123
#usada para o usuario saber de onde vem o link que ele digitou

from Control.notificacao import enviar_notificacao
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from Model.qr_code_bd import qr_bd

def criar_qrcode():

    while True:

        link = input("digite um link (ou escreva sair): ")


        if link.lower() == "sair":
            print("saindo")
            break


        if not (link.startswith("http://") or link.startswith("https://")):
            print("link invalído")
            continue
        #link.startswith("https://") vera se o link é http ou https fazendo assim uma validação
        #se ele realmente é um site valido da internet ou se é apenas uma coisa digitada aleatoriamente pelo usuario


        site = urlparse(link).netloc.replace("www.", "")
        #urlparse ira separar o link em partes para conseguir pegar somente o nome do site
        #replace ira retirar o www para deixar o nome do site mais limpo


        qr = qrcode.make(link)
        #qrcode.make servira para pegar o link utilizado pelo usuario e transformar em qrcode


        nome = f"qr_{int(time.time())}.png"
        #cria um nome único para o arquivo usando o horario atual em segundos
        #evitando que um QR Code sobrescreva o outro


        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        #os.path.expanduser("~") servira para encontrar a pasta do usuario atual
        #os.path.join servira para juntar os caminhos das pastas corretamente


        pasta_qr = os.path.join(desktop, "QR Codes")

        if not os.path.exists(pasta_qr):
            os.makedirs(pasta_qr)

        #os.makedirs ira criar a pasta caso ela ainda não exista


        salvamento = os.path.join(pasta_qr, nome)


        qr.save(salvamento)
        #qr.save servira para salvar o qrcode criado no local escolhido


        salvar_planilha = input(
            "deseja salvar na planilha?(s/n): "
        ).lower()


        if salvar_planilha == "s":


            planilha = os.path.join(
                pasta_qr,
                "QR_Code.xlsx"
            )
            

            if os.path.exists(planilha):

                wb = load_workbook(planilha)
                ws = wb.active
            #verifica se a planilha ja existe para não apagar os QR Codes antigos


            else:
                wb = Workbook()
                ws = wb.active

            ws.title = "Historico QR"
                #criando os nomes das colunas da planilha

            ws["A1"] = "Site"
            ws["B1"] = "Link"
            ws["C1"] = "Data"
            ws["D1"] = "QR Code"


                #define o tamanho das colunas para não ficar tudo apertado

            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 18
            linha = ws.max_row + 1
            #max_row pega a ultima linha usada da planilha
            #assim conseguimos adicionar o novo QR na próxima linha vazia
            ws[f"A{linha}"] = site
            ws[f"B{linha}"] = link
            ws[f"C{linha}"] = datetime.now().strftime("%d/%m/%Y %H:%M")
            img = Image(salvamento)
            #Image ira pegar a imagem do qrcode salva e colocar dentro da planilha
            img.width = 90
            img.height = 90
            ws.add_image(img, f"D{linha}")
            ws.row_dimensions[linha].height = 70
            #aumenta o tamanho da linha para a imagem não ficar por cima dos textos
            wb.save(planilha)
            #salva todas as alterações feitas na planilha
        enviar_notificacao(
            "QR Code",
            "QR Code gerado com sucesso!"
        )


        print("QR Code salvo com sucesso!")

        qr_bd(link, salvamento)
    